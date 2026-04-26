"""
Router de administración.
Dashboard, usuarios CRUD, horarios, dias-no-laborables, reportes y jobs manuales.
"""
import re
from datetime import date
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy.orm import Session


def _norm_grado(v: str) -> str:
    return re.sub(r'(?i)(er|ro|do|to|vo|mo|th|st|nd)$', '', v.strip()).strip()


def _norm_seccion(v: str) -> str:
    return v.strip().upper()

from core.dependencies import get_current_user, get_db, require_roles
from core.security import get_password_hash
from models.asistencia import Asistencia, Horario
from models.comunicado import ComunicadoDestinatario
from models.dia_no_laborable import DiasNoLaborables, ReporteSemanal
from models.estudiante import Estudiante, ApoderadoEstudiante
from models.horario_excepcion import HorarioExcepcion
from models.usuario import TutorAula, Usuario
from schemas.dia_no_laborable import (
    DiasNoLaborablesCreate,
    DiasNoLaborablesResponse,
    HorarioExcepcionCreate,
    HorarioExcepcionResponse,
    HorarioResponse,
    HorarioUpdate,
)
from schemas.horario_curso import HorarioCursoBulk

router = APIRouter()


# ---------------------------------------------------------------------------
# GET /admin/dashboard
# ---------------------------------------------------------------------------

@router.get("/dashboard")
def dashboard(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    hoy = date.today()

    total_estudiantes = db.query(Estudiante).filter(Estudiante.activo == True).count()
    total_usuarios = db.query(Usuario).filter(Usuario.activo == True).count()

    # Asistencias de hoy
    asistencias_hoy = db.query(Asistencia).filter(Asistencia.fecha == hoy).count()
    faltas_hoy = (
        db.query(Asistencia)
        .filter(Asistencia.fecha == hoy, Asistencia.estado == "falta")
        .count()
    )
    tardanzas_hoy = (
        db.query(Asistencia)
        .filter(Asistencia.fecha == hoy, Asistencia.estado == "tardanza")
        .count()
    )

    # Comunicados no leídos (total)
    comunicados_no_leidos = (
        db.query(ComunicadoDestinatario)
        .filter(ComunicadoDestinatario.leido_apoderado == False)
        .count()
    )

    # Estudiantes por nivel
    niveles = {}
    for nivel in ("inicial", "primaria", "secundaria"):
        niveles[nivel] = (
            db.query(Estudiante)
            .filter(Estudiante.nivel == nivel, Estudiante.activo == True)
            .count()
        )

    return {
        "total_estudiantes": total_estudiantes,
        "total_usuarios": total_usuarios,
        "hoy": hoy.isoformat(),
        "asistencias_hoy": asistencias_hoy,
        "faltas_hoy": faltas_hoy,
        "tardanzas_hoy": tardanzas_hoy,
        "comunicados_no_leidos": comunicados_no_leidos,
        "estudiantes_por_nivel": niveles,
    }


# ---------------------------------------------------------------------------
# USUARIOS
# ---------------------------------------------------------------------------

@router.get("/usuarios")
def listar_usuarios(
    rol: Optional[str] = Query(None),
    activo: Optional[bool] = Query(None),
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(50, le=200),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    offset = (pagina - 1) * por_pagina
    q = db.query(Usuario)
    if rol:
        q = q.filter(Usuario.rol == rol)
    if activo is not None:
        q = q.filter(Usuario.activo == activo)

    total = q.count()
    usuarios = q.order_by(Usuario.apellido, Usuario.nombre).offset(offset).limit(por_pagina).all()

    return {
        "total": total,
        "pagina": pagina,
        "por_pagina": por_pagina,
        "items": [_usuario_dict(u, db) for u in usuarios],
    }


@router.get("/usuarios/buscar-dni/{dni}")
def buscar_usuario_por_dni(
    dni: str,
    rol: Optional[str] = Query(None, description="Filtrar por rol, e.g. apoderado"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    q = db.query(Usuario).filter(Usuario.dni == dni)
    if rol:
        q = q.filter(Usuario.rol == rol)
    u = q.first()
    if not u:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No se encontró usuario con ese DNI")
    return _usuario_dict(u, db)


@router.get("/usuarios/{usuario_id}")
def obtener_usuario(
    usuario_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    u = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Usuario no encontrado")
    return _usuario_dict(u, db)


@router.post("/usuarios", status_code=201)
def crear_usuario(
    data: dict,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    """
    Crea un usuario. Body:
    {dni, nombre, apellido, email, password, rol, nivel?, nivel_tutor?, grado?, seccion?, anio?}
    """
    campos = ("dni", "nombre", "apellido", "email", "password", "rol")
    for c in campos:
        if not data.get(c):
            raise HTTPException(
                status.HTTP_422_UNPROCESSABLE_ENTITY, f"Campo requerido: {c}"
            )

    if db.query(Usuario).filter(Usuario.email == data["email"]).first():
        raise HTTPException(status.HTTP_409_CONFLICT, "El email ya está registrado")
    if db.query(Usuario).filter(Usuario.dni == data["dni"]).first():
        raise HTTPException(status.HTTP_409_CONFLICT, "El DNI ya está registrado")

    u = Usuario(
        dni=data["dni"],
        nombre=data["nombre"],
        apellido=data["apellido"],
        email=data["email"],
        password_hash=get_password_hash(data["password"]),
        rol=data["rol"],
        nivel=data.get("nivel"),
        telefono=data.get("telefono") or None,
        activo=True,
    )
    db.add(u)
    db.flush()

    # Si es tutor, asignar aula
    if data["rol"] == "tutor" and data.get("grado") and data.get("seccion"):
        nivel_tutor = data.get("nivel_tutor") or data.get("nivel")
        if not nivel_tutor:
            raise HTTPException(
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "Se requiere nivel_tutor para asignar aula",
            )
        db.add(TutorAula(
            tutor_id=u.id,
            nivel=nivel_tutor,
            grado=_norm_grado(data["grado"]),
            seccion=_norm_seccion(data["seccion"]),
            anio=data.get("anio", date.today().year),
        ))
        u.nivel = nivel_tutor  # también en usuarios para acceso rápido por token

    db.commit()
    db.refresh(u)
    return _usuario_dict(u, db)


@router.put("/usuarios/{usuario_id}")
def actualizar_usuario(
    usuario_id: str,
    data: dict,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    u = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Usuario no encontrado")

    campos_editables = ("nombre", "apellido", "email", "dni", "rol", "nivel", "activo", "foto_url", "telefono", "es_apoderado")
    for campo in campos_editables:
        if campo in data:
            # Verificar unicidad de email y dni
            if campo == "email" and data["email"] != u.email:
                if db.query(Usuario).filter(Usuario.email == data["email"]).first():
                    raise HTTPException(status.HTTP_409_CONFLICT, "El email ya está en uso")
            if campo == "dni" and data["dni"] != u.dni:
                if db.query(Usuario).filter(Usuario.dni == data["dni"]).first():
                    raise HTTPException(status.HTTP_409_CONFLICT, "El DNI ya está en uso")
            setattr(u, campo, data[campo])

    # Actualizar aula de tutor si se envían datos
    if u.rol == "tutor" and any(k in data for k in ("grado", "seccion", "nivel_tutor")):
        vinculo = db.query(TutorAula).filter(TutorAula.tutor_id == u.id).first()
        if vinculo:
            if "grado" in data:
                vinculo.grado = _norm_grado(data["grado"])
            if "seccion" in data:
                vinculo.seccion = _norm_seccion(data["seccion"])
            if "nivel_tutor" in data:
                vinculo.nivel = data["nivel_tutor"]
                u.nivel = data["nivel_tutor"]
            if "anio" in data:
                vinculo.anio = data["anio"]
        elif data.get("grado") and data.get("seccion"):
            nivel_t = data.get("nivel_tutor") or u.nivel
            db.add(TutorAula(
                tutor_id=u.id,
                nivel=nivel_t,
                grado=_norm_grado(data["grado"]),
                seccion=_norm_seccion(data["seccion"]),
                anio=data.get("anio", date.today().year),
            ))

    db.commit()
    db.refresh(u)
    return _usuario_dict(u, db)


@router.put("/usuarios/{usuario_id}/password", status_code=204)
def cambiar_password_usuario(
    usuario_id: str,
    data: dict,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    u = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Usuario no encontrado")
    nueva = data.get("nueva_password", "")
    if len(nueva) < 8:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY, "La contraseña debe tener al menos 8 caracteres"
        )
    u.password_hash = get_password_hash(nueva)
    db.commit()


@router.delete("/usuarios/{usuario_id}", status_code=204)
def desactivar_usuario(
    usuario_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    """Desactiva (soft delete) un usuario."""
    if usuario_id == current_user.id:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "No puedes desactivarte a ti mismo")
    u = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not u:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Usuario no encontrado")
    u.activo = False
    db.commit()


def _get_hijos_apoderado(apoderado_id: str, db: Session) -> list:
    vinculos = db.query(ApoderadoEstudiante).filter(
        ApoderadoEstudiante.apoderado_id == apoderado_id
    ).all()
    hijos = []
    for v in vinculos:
        est = db.query(Estudiante).filter(Estudiante.id == v.estudiante_id).first()
        if est:
            hijos.append({
                "id": est.id,
                "nombre": est.nombre,
                "apellido": est.apellido,
                "dni": est.dni,
                "nivel": est.nivel,
                "grado": est.grado,
                "seccion": est.seccion,
                "activo": est.activo,
            })
    return hijos


def _usuario_dict(u: Usuario, db: Session) -> dict:
    d = {
        "id": u.id,
        "dni": u.dni,
        "nombre": u.nombre,
        "apellido": u.apellido,
        "email": u.email,
        "rol": u.rol,
        "nivel": u.nivel,
        "telefono": u.telefono,
        "activo": u.activo,
        "foto_url": u.foto_url,
        "es_apoderado": bool(u.es_apoderado),
        "created_at": u.created_at.isoformat() if u.created_at else None,
        "aula": None,
        "hijos": [],
    }
    if u.rol == "tutor":
        aula = db.query(TutorAula).filter(TutorAula.tutor_id == u.id).first()
        if aula:
            d["aula"] = {
                "nivel": aula.nivel,
                "grado": aula.grado,
                "seccion": aula.seccion,
                "anio": aula.anio,
            }
    if u.es_apoderado or u.rol == "apoderado":
        d["hijos"] = _get_hijos_apoderado(u.id, db)
    return d


# ---------------------------------------------------------------------------
# APODERADOS — gestión de vínculos apoderado-estudiante
# ---------------------------------------------------------------------------

@router.get("/apoderados")
def listar_apoderados(
    q: Optional[str] = Query(None, description="Buscar por nombre, apellido o DNI"),
    activo: Optional[bool] = Query(None),
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(50, le=200),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    offset = (pagina - 1) * por_pagina

    if q:
        # Con búsqueda: incluir TODO el personal activo para poder vincular
        # a cualquier trabajador del colegio que también sea padre/madre
        like = f"%{q}%"
        query = db.query(Usuario).filter(
            (Usuario.nombre.ilike(like)) |
            (Usuario.apellido.ilike(like)) |
            (Usuario.dni.ilike(like))
        )
    else:
        # Sin búsqueda (listado general): solo apoderados reconocidos
        query = db.query(Usuario).filter(
            (Usuario.rol == "apoderado") | (Usuario.es_apoderado == True)
        )

    if activo is not None:
        query = query.filter(Usuario.activo == activo)

    total = query.count()
    apoderados = query.order_by(Usuario.apellido, Usuario.nombre).offset(offset).limit(por_pagina).all()

    items = [{**_usuario_dict(u, db), "hijos": _get_hijos_apoderado(u.id, db)} for u in apoderados]

    return {"total": total, "pagina": pagina, "por_pagina": por_pagina, "items": items}


@router.post("/apoderados/{apoderado_id}/hijos/{estudiante_id}", status_code=201)
def vincular_hijo(
    apoderado_id: str,
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    u = db.query(Usuario).filter(Usuario.id == apoderado_id).first()
    if not u or (u.rol != "apoderado" and not u.es_apoderado):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Apoderado no encontrado")
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")

    existente = db.query(ApoderadoEstudiante).filter(
        ApoderadoEstudiante.apoderado_id == apoderado_id,
        ApoderadoEstudiante.estudiante_id == estudiante_id,
    ).first()
    if existente:
        raise HTTPException(status.HTTP_409_CONFLICT, "El vínculo ya existe")

    db.add(ApoderadoEstudiante(apoderado_id=apoderado_id, estudiante_id=estudiante_id))
    db.commit()
    return {"detail": "Vínculo creado"}


@router.delete("/apoderados/{apoderado_id}/hijos/{estudiante_id}", status_code=204)
def desvincular_hijo(
    apoderado_id: str,
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    vinculo = db.query(ApoderadoEstudiante).filter(
        ApoderadoEstudiante.apoderado_id == apoderado_id,
        ApoderadoEstudiante.estudiante_id == estudiante_id,
    ).first()
    if not vinculo:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Vínculo no encontrado")
    db.delete(vinculo)
    db.commit()


# ---------------------------------------------------------------------------
# APODERADOS — importación masiva Excel
# ---------------------------------------------------------------------------

@router.get("/apoderados/plantilla-excel")
def plantilla_apoderados(
    current_user: Usuario = Depends(require_roles("admin")),
):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apoderados"

    fill_hdr = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font_hdr = Font(color="FFFFFF", bold=True)
    headers  = ["DNI", "Nombre", "Apellido", "Email", "Telefono"]
    widths   = [12, 20, 25, 30, 14]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = Alignment(horizontal="center")
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws_n = wb.create_sheet("Instrucciones")
    fill_t = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    instrucciones = [
        ["Campo",     "Qué escribir"],
        ["DNI",       "8 dígitos numéricos  (ej: 12345678)"],
        ["Nombre",    "Nombre(s) del apoderado"],
        ["Apellido",  "Apellido paterno y materno"],
        ["Email",     "Correo electrónico único — se usa para iniciar sesión"],
        ["Telefono",  "Número de 9 dígitos sin prefijo  (ej: 987654321)  — opcional"],
        ["", ""],
        ["NOTA", "La contraseña inicial será el DNI del apoderado. Puede cambiarla después."],
    ]
    for i, fila in enumerate(instrucciones, 1):
        for col, val in enumerate(fila, 1):
            cell = ws_n.cell(row=i, column=col, value=val)
            if i == 1:
                cell.fill = fill_t
                cell.font = Font(color="FFFFFF", bold=True)
    ws_n.column_dimensions["A"].width = 12
    ws_n.column_dimensions["B"].width = 55

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_apoderados.xlsx"},
    )


@router.post("/apoderados/importar-excel")
async def importar_apoderados_excel(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    from io import BytesIO
    import openpyxl

    ext = (archivo.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Solo se aceptan archivos .xlsx o .xls")

    contenido = await archivo.read()
    if not contenido:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo está vacío")
    if len(contenido) > 5 * 1024 * 1024:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo supera 5 MB")

    try:
        wb  = openpyxl.load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        ws  = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No se pudo leer el archivo Excel")

    if len(rows) < 2:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo no tiene datos")

    raw_headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    ALIASES = {
        "dni":      ["dni"],
        "nombre":   ["nombre"],
        "apellido": ["apellido"],
        "email":    ["email", "correo"],
        "telefono": ["telefono", "teléfono", "cel", "celular"],
    }
    col_idx: dict = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            if alias in raw_headers:
                col_idx[field] = raw_headers.index(alias)
                break

    missing = [f for f in ("dni", "nombre", "apellido", "email") if f not in col_idx]
    if missing:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
            f"Columnas faltantes: {', '.join(missing)}. Descarga la plantilla.")

    def _get(row, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row): return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    importados = 0
    omitidos   = 0
    errores    = []
    dni_batch  = {}

    for row_num, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        dni      = _get(row, "dni")
        nombre   = _get(row, "nombre")
        apellido = _get(row, "apellido")
        email    = _get(row, "email").lower()
        telefono = _get(row, "telefono") or None

        def _err(msg):
            errores.append({"fila": row_num, **({"dni": dni} if dni else {}), "motivo": msg})

        if not dni:      _err("DNI vacío");      omitidos += 1; continue
        if not nombre:   _err("Nombre vacío");   omitidos += 1; continue
        if not apellido: _err("Apellido vacío"); omitidos += 1; continue
        if not email:    _err("Email vacío");    omitidos += 1; continue

        # Normalizar DNI (Excel puede traer float)
        if "." in dni:
            dni = dni.split(".")[0]
        if not dni.isdigit() or len(dni) != 8:
            _err(f"DNI inválido '{dni}' (debe ser 8 dígitos)"); omitidos += 1; continue

        if "@" not in email or "." not in email.split("@")[-1]:
            _err(f"Email inválido '{email}'"); omitidos += 1; continue

        # Teléfono: limpiar y normalizar
        if telefono:
            tel_num = re.sub(r'[^0-9]', '', telefono.split(".")[0])
            telefono = tel_num if tel_num else None

        if dni in dni_batch:
            _err(f"DNI duplicado en el archivo (fila {dni_batch[dni]})"); omitidos += 1; continue
        dni_batch[dni] = row_num

        if db.query(Usuario).filter(Usuario.dni == dni).first():
            _err(f"DNI ya registrado"); omitidos += 1; continue
        if db.query(Usuario).filter(Usuario.email == email).first():
            _err(f"Email ya registrado"); omitidos += 1; continue

        try:
            db.add(Usuario(
                dni=dni,
                nombre=nombre,
                apellido=apellido,
                email=email,
                password_hash=get_password_hash(dni),
                rol="apoderado",
                telefono=telefono,
                activo=True,
            ))
            db.flush()
            importados += 1
        except Exception as exc:
            _err(f"Error al crear: {exc}"); omitidos += 1

    if importados:
        db.commit()
    else:
        db.rollback()

    return {"importados": importados, "omitidos": omitidos, "errores": errores}


# ---------------------------------------------------------------------------
# HORARIOS
# ---------------------------------------------------------------------------

_ROL_TO_NIVEL  = {"i-auxiliar": "inicial", "p-auxiliar": "primaria", "s-auxiliar": "secundaria"}
_ROLES_HORARIO = ("admin", "i-auxiliar", "p-auxiliar", "s-auxiliar", "directivo")

@router.get("/horarios", response_model=List[HorarioResponse])
def listar_horarios(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(*_ROLES_HORARIO)),
):
    q = db.query(Horario)
    if current_user.rol in _ROL_TO_NIVEL:
        q = q.filter(Horario.nivel == _ROL_TO_NIVEL[current_user.rol])
    return q.order_by(Horario.nivel).all()


@router.put("/horarios/{nivel}", response_model=HorarioResponse)
def actualizar_horario(
    nivel: str,
    data: HorarioUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(*_ROLES_HORARIO)),
):
    if nivel not in ("inicial", "primaria", "secundaria"):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Nivel inválido")
    if current_user.rol in _ROL_TO_NIVEL and nivel != _ROL_TO_NIVEL[current_user.rol]:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Solo puedes editar el horario de tu nivel")

    horario = db.query(Horario).filter(Horario.nivel == nivel).first()
    if not horario:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Horario no encontrado")

    for campo, valor in data.model_dump(exclude_none=True).items():
        setattr(horario, campo, valor)

    db.commit()
    db.refresh(horario)

    # Reagendar el job de faltas con la nueva hora_cierre_faltas
    try:
        from services.scheduler import reschedule_faltas
        reschedule_faltas(nivel, horario.hora_cierre_faltas)
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("[Scheduler] No se pudo reagendar faltas_%s: %s", nivel, exc)

    return horario


# ---------------------------------------------------------------------------
# EXCEPCIONES DE HORARIO
# ---------------------------------------------------------------------------

NIVELES_VALIDOS = {"inicial", "primaria", "secundaria", "todos"}


@router.get("/horarios/excepciones", response_model=List[HorarioExcepcionResponse])
def listar_excepciones(
    desde: Optional[date] = Query(None, description="Filtrar desde esta fecha (default: hoy)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(*_ROLES_HORARIO)),
):
    """Devuelve las excepciones de horario desde `desde` en adelante."""
    fecha_min = desde or date.today()
    q = db.query(HorarioExcepcion).filter(HorarioExcepcion.fecha >= fecha_min)
    if current_user.rol in _ROL_TO_NIVEL:
        nivel_aux = _ROL_TO_NIVEL[current_user.rol]
        q = q.filter(HorarioExcepcion.nivel.in_([nivel_aux, "todos"]))
    return q.order_by(HorarioExcepcion.fecha, HorarioExcepcion.nivel).all()


@router.post("/horarios/excepciones", response_model=HorarioExcepcionResponse, status_code=201)
def crear_excepcion(
    data: HorarioExcepcionCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(*_ROLES_HORARIO)),
):
    """Crea una excepción de horario para una fecha y nivel específicos."""
    # Auxiliares solo pueden crear excepciones para su propio nivel
    if current_user.rol in _ROL_TO_NIVEL:
        data.nivel = _ROL_TO_NIVEL[current_user.rol]
    if data.nivel not in NIVELES_VALIDOS:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Nivel inválido")

    if not data.hora_ingreso_fin and not data.hora_salida_inicio and not data.hora_cierre_faltas:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "Debes modificar al menos un horario (ingreso, salida o cierre de faltas)",
        )

    existente = db.query(HorarioExcepcion).filter(
        HorarioExcepcion.nivel == data.nivel,
        HorarioExcepcion.fecha  == data.fecha,
    ).first()
    if existente:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Ya existe una excepción para '{data.nivel}' el {data.fecha}. Elimínala primero para reemplazarla.",
        )

    nueva = HorarioExcepcion(
        nivel=data.nivel,
        fecha=data.fecha,
        hora_ingreso_fin=data.hora_ingreso_fin,
        hora_salida_inicio=data.hora_salida_inicio,
        hora_cierre_faltas=data.hora_cierre_faltas,
        motivo=data.motivo,
        created_by=current_user.id,
    )
    db.add(nueva)
    db.commit()
    db.refresh(nueva)

    # Push de aviso de cambio de horario a apoderados del nivel
    from services.firebase_service import push_aviso_horario_bg
    push_aviso_horario_bg(
        nivel=data.nivel,
        fecha=data.fecha,
        hora_ingreso_fin=data.hora_ingreso_fin,
        hora_salida_inicio=data.hora_salida_inicio,
        hora_cierre_faltas=data.hora_cierre_faltas,
        motivo=data.motivo,
    )

    return nueva


@router.delete("/horarios/excepciones/{excepcion_id}", status_code=204)
def eliminar_excepcion(
    excepcion_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(*_ROLES_HORARIO)),
):
    exc = db.query(HorarioExcepcion).filter(HorarioExcepcion.id == excepcion_id).first()
    if not exc:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Excepción no encontrada")
    if current_user.rol in _ROL_TO_NIVEL and exc.nivel != _ROL_TO_NIVEL[current_user.rol]:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "No puedes eliminar excepciones de otro nivel")
    db.delete(exc)
    db.commit()


# ---------------------------------------------------------------------------
# DÍAS NO LABORABLES
# ---------------------------------------------------------------------------

@router.get("/dias-no-laborables", response_model=List[DiasNoLaborablesResponse])
def listar_dias_no_laborables(
    anio: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    anio_filtro = anio or date.today().year
    return (
        db.query(DiasNoLaborables)
        .filter(
            DiasNoLaborables.fecha >= date(anio_filtro, 1, 1),
            DiasNoLaborables.fecha <= date(anio_filtro, 12, 31),
        )
        .order_by(DiasNoLaborables.fecha.asc())
        .all()
    )


@router.post("/dias-no-laborables", status_code=201)
def crear_dia_no_laborable(
    data: DiasNoLaborablesCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    from datetime import timedelta

    fecha_fin = data.fecha_fin or data.fecha_inicio
    if fecha_fin < data.fecha_inicio:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "fecha_fin no puede ser anterior a fecha_inicio")

    creados = []
    omitidos = 0
    cursor = data.fecha_inicio

    while cursor <= fecha_fin:
        existente = db.query(DiasNoLaborables).filter(DiasNoLaborables.fecha == cursor).first()
        if existente:
            omitidos += 1
        else:
            dia = DiasNoLaborables(
                fecha=cursor,
                tipo=data.tipo,
                nivel=data.nivel or "todos",
                grado=data.grado or None,
                seccion=data.seccion or None,
                motivo=data.descripcion,
                created_by=current_user.id,
            )
            db.add(dia)
            creados.append(cursor)
        cursor += timedelta(days=1)

    db.commit()

    total = len(creados)
    if total == 0:
        raise HTTPException(status.HTTP_409_CONFLICT, f"Todos los días del rango ya estaban marcados ({omitidos} omitidos)")

    # Push a apoderados si el admin lo solicitó
    if data.notificar and total > 0:
        from services.firebase_service import push_aviso_calendario_bg
        push_aviso_calendario_bg(
            nivel=data.nivel,
            grado=data.grado,
            seccion=data.seccion,
            fecha_inicio=data.fecha_inicio,
            fecha_fin=data.fecha_fin or data.fecha_inicio,
            motivo=data.descripcion,
            tipo=data.tipo,
        )

    return {"creados": total, "omitidos": omitidos}


@router.delete("/dias-no-laborables/{dia_id}", status_code=204)
def eliminar_dia_no_laborable(
    dia_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    dia = db.query(DiasNoLaborables).filter(DiasNoLaborables.id == dia_id).first()
    if not dia:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Día no laborable no encontrado")
    db.delete(dia)
    db.commit()


# ---------------------------------------------------------------------------
# REPORTES
# ---------------------------------------------------------------------------

@router.get("/reportes")
def listar_reportes(
    nivel: Optional[str] = Query(None),
    semana_inicio: Optional[date] = Query(None),
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(50, le=200),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    offset = (pagina - 1) * por_pagina

    q = (
        db.query(ReporteSemanal)
        .join(Estudiante, Estudiante.id == ReporteSemanal.estudiante_id)
    )
    if nivel:
        q = q.filter(Estudiante.nivel == nivel)
    if semana_inicio:
        q = q.filter(ReporteSemanal.semana_inicio == semana_inicio)

    total = q.count()
    reportes = (
        q.order_by(ReporteSemanal.semana_inicio.desc(), Estudiante.apellido)
        .offset(offset)
        .limit(por_pagina)
        .all()
    )

    items = []
    for r in reportes:
        est = db.query(Estudiante).filter(Estudiante.id == r.estudiante_id).first()
        items.append({
            "id": r.id,
            "semana_inicio": r.semana_inicio.isoformat(),
            "semana_fin": r.semana_fin.isoformat(),
            "dias_asistio": r.dias_asistio,
            "dias_laborables": r.dias_laborables,
            "tardanzas": r.tardanzas,
            "faltas": r.faltas,
            "comunicados_pendientes": r.comunicados_pendientes,
            "correo_enviado": r.correo_enviado,
            "estudiante": {
                "id": est.id,
                "nombre": est.nombre,
                "apellido": est.apellido,
                "nivel": est.nivel,
                "grado": est.grado,
                "seccion": est.seccion,
            } if est else None,
        })

    return {"total": total, "pagina": pagina, "por_pagina": por_pagina, "items": items}


# ---------------------------------------------------------------------------
# ASISTENCIA — Vista general por fecha
# ---------------------------------------------------------------------------

@router.get("/asistencia")
def asistencia_general(
    fecha: Optional[date] = Query(None),
    nivel: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    hoy = fecha or date.today()

    q = (
        db.query(Asistencia)
        .join(Estudiante, Estudiante.id == Asistencia.estudiante_id)
        .filter(Asistencia.fecha == hoy)
    )
    if nivel:
        q = q.filter(Estudiante.nivel == nivel)

    registros = q.order_by(Asistencia.hora.asc()).all()

    result = []
    for a in registros:
        est = db.query(Estudiante).filter(Estudiante.id == a.estudiante_id).first()
        result.append({
            "id": a.id,
            "fecha": a.fecha.isoformat(),
            "tipo": a.tipo,
            "estado": a.estado,
            "hora": a.hora.isoformat() if a.hora else None,
            "observacion": a.observacion,
            "estudiante": {
                "id": est.id,
                "nombre": est.nombre,
                "apellido": est.apellido,
                "nivel": est.nivel,
                "grado": est.grado,
                "seccion": est.seccion,
            } if est else None,
        })

    return {"fecha": hoy.isoformat(), "total": len(result), "registros": result}


# ---------------------------------------------------------------------------
# Jobs manuales
# ---------------------------------------------------------------------------

@router.post("/jobs/faltas/{nivel}", status_code=202)
def trigger_faltas(
    nivel: str,
    current_user: Usuario = Depends(require_roles("admin")),
):
    if nivel not in ("inicial", "primaria", "secundaria"):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Nivel inválido")

    from services.scheduler import registrar_faltas_nivel
    import threading

    threading.Thread(target=registrar_faltas_nivel, args=(nivel,), daemon=True).start()
    return {"detail": f"Job faltas_{nivel} disparado en background"}


@router.post("/jobs/reporte-semanal", status_code=202)
def trigger_reporte(
    current_user: Usuario = Depends(require_roles("admin")),
):
    from services.scheduler import generar_reportes_semanales
    import threading

    threading.Thread(target=generar_reportes_semanales, daemon=True).start()
    return {"detail": "Job reporte_semanal disparado en background"}


# ---------------------------------------------------------------------------
# HORARIO DE CLASES — archivo (PDF / imagen) por aula
# ---------------------------------------------------------------------------

@router.get("/horario-clases")
def get_horario_archivo(
    nivel:   str           = Query(...),
    grado:   str           = Query(...),
    seccion: str           = Query(...),
    anio:    Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(*_ROLES_HORARIO)),
):
    if current_user.rol in _ROL_TO_NIVEL and nivel != _ROL_TO_NIVEL[current_user.rol]:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Solo puedes ver horarios de tu nivel")
    """Devuelve el archivo de horario de clases para un aula, si existe."""
    from models.horario_archivo import HorarioArchivo
    year = anio or date.today().year
    h = db.query(HorarioArchivo).filter(
        HorarioArchivo.nivel   == nivel,
        HorarioArchivo.grado   == grado,
        HorarioArchivo.seccion == seccion.upper(),
        HorarioArchivo.anio    == year,
    ).first()
    if not h:
        return None
    return {
        "id":             str(h.id),
        "archivo_nombre": h.archivo_nombre,
        "archivo_url":    h.archivo_url,
        "created_at":     h.created_at.isoformat() if h.created_at else None,
    }


@router.post("/horario-clases/subir", status_code=201)
async def subir_horario_archivo(
    nivel:   str        = Query(...),
    grado:   str        = Query(...),
    seccion: str        = Query(...),
    anio:    int        = Query(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(*_ROLES_HORARIO)),
):
    if current_user.rol in _ROL_TO_NIVEL and nivel != _ROL_TO_NIVEL[current_user.rol]:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Solo puedes subir horarios de tu nivel")
    """Sube (o reemplaza) el horario de clases de un aula como PDF o imagen."""
    from models.horario_archivo import HorarioArchivo
    from services.drive_service import subir_archivo

    contenido = await archivo.read()
    mime = archivo.content_type or "application/octet-stream"
    try:
        resultado = subir_archivo(contenido, archivo.filename, mime)
    except (ValueError, RuntimeError) as e:
        raise HTTPException(status.HTTP_503_SERVICE_UNAVAILABLE, str(e))

    sec = seccion.upper()

    # Si ya existe uno para este aula/año, lo reemplaza
    existente = db.query(HorarioArchivo).filter(
        HorarioArchivo.nivel   == nivel,
        HorarioArchivo.grado   == grado,
        HorarioArchivo.seccion == sec,
        HorarioArchivo.anio    == anio,
    ).first()

    if existente:
        existente.archivo_nombre = archivo.filename
        existente.archivo_url    = resultado["url"]
        existente.subido_por     = str(current_user.id)
    else:
        nuevo = HorarioArchivo(
            nivel=nivel, grado=grado, seccion=sec, anio=anio,
            archivo_nombre=archivo.filename,
            archivo_url=resultado["url"],
            subido_por=str(current_user.id),
        )
        db.add(nuevo)

    db.commit()
    return {"detail": "Horario subido correctamente", "url": resultado["url"]}


@router.delete("/horario-clases/{horario_id}", status_code=204)
def eliminar_horario_archivo(
    horario_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles(*_ROLES_HORARIO)),
):
    from models.horario_archivo import HorarioArchivo
    h = db.query(HorarioArchivo).filter(HorarioArchivo.id == horario_id).first()
    if not h:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Horario no encontrado")
    if current_user.rol in _ROL_TO_NIVEL and h.nivel != _ROL_TO_NIVEL[current_user.rol]:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Solo puedes eliminar horarios de tu nivel")
    db.delete(h)
    db.commit()


# ---------------------------------------------------------------------------
# HORARIO DE CURSOS (horario semanal de clases por aula)
# ---------------------------------------------------------------------------

@router.get("/horarios-curso")
def listar_horarios_curso(
    nivel:   Optional[str] = Query(None),
    grado:   Optional[str] = Query(None),
    seccion: Optional[str] = Query(None),
    anio:    Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    """Lista periodos del horario de clases, opcionalmente filtrados por aula/año."""
    from models.horario_curso import HorarioCurso
    q = db.query(HorarioCurso)
    if nivel:   q = q.filter(HorarioCurso.nivel   == nivel)
    if grado:   q = q.filter(HorarioCurso.grado   == grado)
    if seccion: q = q.filter(HorarioCurso.seccion == seccion.upper())
    if anio:    q = q.filter(HorarioCurso.anio    == anio)
    periodos = q.order_by(HorarioCurso.dia_semana, HorarioCurso.hora_inicio).all()
    return [
        {
            "id":             str(p.id),
            "nivel":          p.nivel,
            "grado":          p.grado,
            "seccion":        p.seccion,
            "anio":           p.anio,
            "dia_semana":     p.dia_semana,
            "hora_inicio":    p.hora_inicio,
            "hora_fin":       p.hora_fin,
            "curso_nombre":   p.curso_nombre,
            "docente_nombre": p.docente_nombre,
        }
        for p in periodos
    ]


@router.post("/horarios-curso", status_code=201)
def cargar_horario_curso(
    payload: HorarioCursoBulk,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    """
    Reemplaza el horario completo de un aula para un año.
    Si ya existe horario para ese nivel/grado/sección/año, lo elimina primero.
    """
    from models.horario_curso import HorarioCurso

    seccion = payload.seccion.upper()

    # Limpiar horario previo del aula/año
    db.query(HorarioCurso).filter(
        HorarioCurso.nivel   == payload.nivel,
        HorarioCurso.grado   == payload.grado,
        HorarioCurso.seccion == seccion,
        HorarioCurso.anio    == payload.anio,
    ).delete()

    nuevos = [
        HorarioCurso(
            nivel=payload.nivel,
            grado=payload.grado,
            seccion=seccion,
            anio=payload.anio,
            dia_semana=p.dia_semana,
            hora_inicio=p.hora_inicio,
            hora_fin=p.hora_fin,
            curso_nombre=p.curso_nombre,
            docente_nombre=p.docente_nombre,
        )
        for p in payload.periodos
    ]
    db.add_all(nuevos)
    db.commit()
    return {"creados": len(nuevos), "aula": f"{payload.nivel} {payload.grado}° {seccion}"}


@router.delete("/horarios-curso/{periodo_id}", status_code=204)
def eliminar_periodo_horario(
    periodo_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    """Elimina un periodo individual del horario."""
    from models.horario_curso import HorarioCurso
    p = db.query(HorarioCurso).filter(HorarioCurso.id == periodo_id).first()
    if not p:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Periodo no encontrado")
    db.delete(p)
    db.commit()


@router.get("/jobs")
def listar_jobs(
    current_user: Usuario = Depends(require_roles("admin")),
):
    from services.scheduler import scheduler

    return [
        {
            "id": j.id,
            "next_run": str(j.next_run_time),
            "trigger": str(j.trigger),
        }
        for j in scheduler.get_jobs()
    ]


@router.delete("/asistencia/hoy", status_code=200)
def borrar_asistencia_hoy(
    estudiante_id: Optional[str] = Query(None, description="Si se omite, borra TODOS los registros de hoy"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    """
    Elimina registros de asistencia de HOY (solo admin).
    Útil para pruebas — permite re-escanear el mismo día.
    """
    from models.asistencia import Asistencia
    from core.tz import hoy as _hoy

    fecha_hoy = _hoy()
    q = db.query(Asistencia).filter(Asistencia.fecha == fecha_hoy)
    if estudiante_id:
        q = q.filter(Asistencia.estudiante_id == estudiante_id)

    total = q.count()
    q.delete(synchronize_session=False)
    db.commit()
    return {"eliminados": total, "fecha": str(fecha_hoy), "estudiante_id": estudiante_id or "todos"}
