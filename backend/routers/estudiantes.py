import base64
import re
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from core.dependencies import get_current_user, get_db, require_roles
from core.security import get_password_hash
from models.estudiante import ApoderadoEstudiante, Estudiante
from models.usuario import Usuario
from schemas.estudiante import (
    DesactivarRequest, EstudianteCreate, EstudianteResponse,
    EstudianteUpdate, ReactivarRequest,
)
from services.qr_service import generar_qr_png, generar_qr_solo_png, generar_qr_token

router = APIRouter()


# ---------------------------------------------------------------------------
# GET /estudiantes
# ---------------------------------------------------------------------------

@router.get("/", response_model=List[EstudianteResponse])
def listar(
    nivel: Optional[str] = Query(None),
    grado: Optional[str] = Query(None),
    seccion: Optional[str] = Query(None),
    activo: Optional[bool] = Query(True),
    q: Optional[str] = Query(None, description="Buscar por nombre, apellido o DNI"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    query = db.query(Estudiante)

    if activo is not None:
        query = query.filter(Estudiante.activo == activo)
    if nivel:
        query = query.filter(Estudiante.nivel == nivel)
    if grado:
        query = query.filter(Estudiante.grado == grado)
    if seccion:
        query = query.filter(Estudiante.seccion == seccion)
    if q:
        like = f"%{q}%"
        query = query.filter(
            Estudiante.nombre.ilike(like)
            | Estudiante.apellido.ilike(like)
            | Estudiante.dni.ilike(like)
        )

    return query.order_by(Estudiante.apellido, Estudiante.nombre).all()


# ---------------------------------------------------------------------------
# GET /estudiantes/plantilla-excel  →  Descargar plantilla Excel vacía
# ---------------------------------------------------------------------------

@router.get("/plantilla-excel")
def descargar_plantilla(
    nivel: Optional[str] = Query(None, description="inicial | primaria | secundaria"),
    current_user: Usuario = Depends(require_roles("admin")),
):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    AULAS_INICIAL_T = {
        "2": ["Amarilla"],
        "3": ["Crema", "Blanca"],
        "4": ["Verde Limon", "Rosada"],
        "5": ["Celeste", "Lila"],
    }
    COLORES_HEADER = {
        "inicial": "2D7A3A", "primaria": "1A5EA8", "secundaria": "6B3FA0",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    color = COLORES_HEADER.get(nivel, "1E3A5F")
    fill_hdr = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font_hdr = Font(color="FFFFFF", bold=True)
    aln_c    = Alignment(horizontal="center")

    # Encabezados sin columna Nivel — el nivel se pasa como param al importar
    if nivel == "inicial":
        headers = ["DNI", "Nombre", "Apellido", "Edad", "Aula"]
        widths  = [12, 20, 25, 8, 16]
    else:
        headers = ["DNI", "Nombre", "Apellido", "Grado", "Seccion"]
        widths  = [12, 20, 25, 8, 12]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = aln_c
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # Dropdowns según nivel
    if nivel == "inicial":
        dv_edad = DataValidation(type="list", formula1='"2,3,4,5"', allow_blank=True)
        dv_edad.sqref = "D2:D501"
        ws.add_data_validation(dv_edad)
        todos_colores = [c for cols in AULAS_INICIAL_T.values() for c in cols]
        dv_aula = DataValidation(type="list", formula1='"{}"'.format(",".join(todos_colores)), allow_blank=True)
        dv_aula.sqref = "E2:E501"
        ws.add_data_validation(dv_aula)
    elif nivel == "primaria":
        dv_g = DataValidation(type="list", formula1='"1,2,3,4,5,6"', allow_blank=True)
        dv_g.sqref = "D2:D501"
        ws.add_data_validation(dv_g)
        dv_s = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True)
        dv_s.sqref = "E2:E501"
        ws.add_data_validation(dv_s)
    elif nivel == "secundaria":
        dv_g = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
        dv_g.sqref = "D2:D501"
        ws.add_data_validation(dv_g)
        dv_s = DataValidation(type="list", formula1='"A,B,C,D,E"', allow_blank=True)
        dv_s.sqref = "E2:E501"
        ws.add_data_validation(dv_s)

    # Hoja instrucciones
    ws_n = wb.create_sheet("Instrucciones")
    fill_t = PatternFill(start_color=color, end_color=color, fill_type="solid")
    rows_inst = [["Campo", "Qué escribir"]]
    rows_inst.append(["DNI", "8 dígitos numéricos  (ej: 12345678)"])
    rows_inst.append(["Nombre", "Solo el nombre del estudiante"])
    rows_inst.append(["Apellido", "Apellido paterno y materno"])
    if nivel == "inicial":
        rows_inst.append(["Edad", "2, 3, 4 o 5  (años de edad)"])
        rows_inst.append(["Aula", "Nombre del color del aula:"])
        for g, cols in AULAS_INICIAL_T.items():
            rows_inst.append([f"  · {g} años", "  →  " + "  ó  ".join(cols)])
    elif nivel == "primaria":
        rows_inst.append(["Grado", "1, 2, 3, 4, 5 o 6"])
        rows_inst.append(["Seccion", "A o B  (grados 1°,2°,3°,6°)"])
        rows_inst.append(["Seccion", "A, B o C  (grados 4° y 5°)"])
    elif nivel == "secundaria":
        rows_inst.append(["Grado", "1, 2, 3, 4 o 5"])
        rows_inst.append(["Seccion", "A, B, C, D o E"])

    for i, fila in enumerate(rows_inst, 1):
        for col, val in enumerate(fila, 1):
            cell = ws_n.cell(row=i, column=col, value=val)
            if i == 1:
                cell.fill = fill_t
                cell.font = Font(color="FFFFFF", bold=True)
    ws_n.column_dimensions["A"].width = 20
    ws_n.column_dimensions["B"].width = 42

    nombre_archivo = f"plantilla_{nivel}.xlsx" if nivel else "plantilla_estudiantes.xlsx"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


# ---------------------------------------------------------------------------
# POST /estudiantes/importar-excel  (admin)
# ---------------------------------------------------------------------------

@router.post("/importar-excel")
async def importar_excel(
    archivo: UploadFile = File(...),
    nivel_param: Optional[str] = Query(None, alias="nivel"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    from io import BytesIO

    import openpyxl

    NIVELES_VALIDOS = {"inicial", "primaria", "secundaria"}
    GRADOS_POR_NIVEL_MAP = {
        "inicial":    {"2", "3", "4", "5"},
        "primaria":   {"1", "2", "3", "4", "5", "6"},
        "secundaria": {"1", "2", "3", "4", "5"},
    }
    # Aulas inicial: letra → nombre de color (acepta ambos en el Excel)
    AULAS_INICIAL = {
        "2": ["Amarilla"],
        "3": ["Crema", "Blanca"],
        "4": ["Verde Limon", "Rosada"],
        "5": ["Celeste", "Lila"],
    }
    SECCIONES_PRIMARIA = {
        "1": {"A", "B"},
        "2": {"A", "B"},
        "3": {"A", "B"},
        "4": {"A", "B", "C"},
        "5": {"A", "B", "C"},
        "6": {"A", "B"},
    }
    SECCIONES_SECUNDARIA = {"A", "B", "C", "D", "E"}

    # Validar extensión
    ext = (archivo.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Solo se aceptan archivos .xlsx o .xls")

    contenido = await archivo.read()
    if not contenido:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo está vacío")
    if len(contenido) > 5 * 1024 * 1024:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo supera el límite de 5 MB")

    # Parsear Excel
    try:
        wb = openpyxl.load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx)",
        )

    if len(rows) < 2:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "El archivo no contiene datos (mínimo 1 fila de encabezados + 1 fila de datos)",
        )

    # Mapear columnas por nombre (flexible, acepta tildes y variantes)
    raw_headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    ALIASES = {
        "dni":      ["dni"],
        "nombre":   ["nombre"],
        "apellido": ["apellido"],
        "nivel":    ["nivel"],
        "grado":    ["grado", "edad"],
        "seccion":  ["seccion", "sección", "aula"],
        "foto_url": ["foto_url", "foto url", "foto"],
    }
    col_idx: dict = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            if alias in raw_headers:
                col_idx[field] = raw_headers.index(alias)
                break

    # nivel puede venir como columna en el Excel O como parámetro de la URL
    tiene_col_nivel = "nivel" in col_idx
    if not tiene_col_nivel and not nivel_param:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Falta indicar el nivel (inicial, primaria o secundaria)")

    campos_req = ["dni", "nombre", "apellido", "grado", "seccion"]
    missing = [f for f in campos_req if f not in col_idx]
    if missing:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Columnas requeridas faltantes: {', '.join(missing)}. Descarga la plantilla para ver el formato correcto.",
        )

    # Procesar fila por fila
    importados = 0
    omitidos = 0
    errores: list = []
    dni_batch: dict = {}

    def _get(row: tuple, field: str) -> str:
        idx = col_idx.get(field)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    for row_num, row in enumerate(rows[1:], start=2):
        # Ignorar filas completamente vacías
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        dni       = _get(row, "dni")
        nombre    = _get(row, "nombre")
        apellido  = _get(row, "apellido")
        nivel     = (_get(row, "nivel").lower() if tiene_col_nivel else (nivel_param or "")).lower().strip()
        grado_raw = _get(row, "grado")
        sec_raw   = _get(row, "seccion")
        foto_url  = _get(row, "foto_url") or None

        def _err(motivo: str):
            errores.append({"fila": row_num, **( {"dni": dni} if dni else {}), "motivo": motivo})

        # Campos vacíos
        if not dni:       _err("DNI vacío");       omitidos += 1; continue
        if not nombre:    _err("Nombre vacío");    omitidos += 1; continue
        if not apellido:  _err("Apellido vacío");  omitidos += 1; continue
        if not nivel:     _err("Nivel vacío");     omitidos += 1; continue
        if not grado_raw: _err("Grado vacío");     omitidos += 1; continue
        if not sec_raw:   _err("Sección vacía");   omitidos += 1; continue

        # Excel puede entregar números como float ("12345678.0")
        if "." in dni:
            dni = dni.split(".")[0]
        if not dni.isdigit() or len(dni) != 8:
            _err(f"DNI inválido '{dni}' (debe ser 8 dígitos numéricos)"); omitidos += 1; continue

        if nivel not in NIVELES_VALIDOS:
            _err(f"Nivel inválido '{nivel}' (válidos: inicial, primaria, secundaria)"); omitidos += 1; continue

        # Normalizar grado: extraer solo el número (maneja "2", 2, 2.0, "2to", etc.)
        grado_num = re.sub(r'[^0-9]', '', grado_raw.split(".")[0].strip())
        grado = grado_num if grado_num else grado_raw.strip()
        if grado not in GRADOS_POR_NIVEL_MAP.get(nivel, set()):
            _err(f"Edad/Grado '{grado_raw}' no válido para {nivel} (válidos: {', '.join(sorted(GRADOS_POR_NIVEL_MAP.get(nivel, set())))})")
            omitidos += 1; continue

        # Validar aula/sección según nivel
        if nivel == "inicial":
            colores = AULAS_INICIAL.get(grado, [])
            letras_validas = [chr(65 + i) for i in range(len(colores))]
            # normalizar: quitar espacios extra, comparar sin tildes ni mayúsculas
            sec_norm = " ".join(sec_raw.strip().split())
            if sec_norm.upper() in letras_validas:
                seccion = sec_norm.upper()
            elif sec_norm.lower() in [c.lower() for c in colores]:
                seccion = letras_validas[[c.lower() for c in colores].index(sec_norm.lower())]
            else:
                validas_str = " | ".join(f"{l}={c}" for l, c in zip(letras_validas, colores))
                _err(f"Aula '{sec_raw}' inválida para {grado} años. Válidas: {validas_str}")
                omitidos += 1; continue
        elif nivel == "primaria":
            seccion = sec_raw.strip().upper()
            validas = SECCIONES_PRIMARIA.get(grado, set())
            if seccion not in validas:
                _err(f"Sección '{seccion}' inválida para {grado}° primaria (válidas: {', '.join(sorted(validas))})")
                omitidos += 1; continue
        else:
            seccion = sec_raw.strip().upper()
            if seccion not in SECCIONES_SECUNDARIA:
                _err(f"Sección '{seccion}' inválida (válidas: A, B, C, D, E)")
                omitidos += 1; continue

        # Duplicado dentro del mismo archivo
        if dni in dni_batch:
            _err(f"DNI duplicado en el archivo (primera aparición en fila {dni_batch[dni]})")
            omitidos += 1; continue
        dni_batch[dni] = row_num

        # Duplicado en la base de datos
        existing = db.query(Estudiante).filter(Estudiante.dni == dni).first()
        if existing:
            estado = "activo" if existing.activo else "inactivo"
            _err(f"DNI ya registrado — {existing.nombre} {existing.apellido} ({estado})")
            omitidos += 1; continue

        # Crear el estudiante
        try:
            db.add(Estudiante(
                dni=dni,
                nombre=nombre,
                apellido=apellido,
                nivel=nivel,
                grado=grado,
                seccion=seccion,
                foto_url=foto_url,
                qr_token=generar_qr_token(),
                activo=True,
            ))
            db.flush()
            importados += 1
        except Exception:
            db.rollback()
            _err("Error al guardar en la base de datos")
            omitidos += 1

    db.commit()
    return {"importados": importados, "omitidos": omitidos, "errores": errores}


# ---------------------------------------------------------------------------
# GET /estudiantes/{id}
# ---------------------------------------------------------------------------

@router.get("/{estudiante_id}", response_model=EstudianteResponse)
def obtener(
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")
    return est


# ---------------------------------------------------------------------------
# POST /estudiantes/{id}/foto  →  Subir / reemplazar foto del estudiante
# ---------------------------------------------------------------------------

@router.post("/{estudiante_id}/foto")
async def subir_foto(
    estudiante_id: str,
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    import base64 as b64lib
    from io import BytesIO
    from PIL import Image

    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")

    if not archivo.content_type or not archivo.content_type.startswith("image/"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Solo se aceptan imágenes (jpg, png, webp...)")

    contenido = await archivo.read()
    if len(contenido) > 5 * 1024 * 1024:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "La imagen supera el límite de 5 MB")

    try:
        img = Image.open(BytesIO(contenido)).convert("RGB")
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No se pudo leer la imagen")

    # Recorte cuadrado centrado
    w, h = img.size
    min_dim = min(w, h)
    left = (w - min_dim) // 2
    top  = (h - min_dim) // 2
    img  = img.crop((left, top, left + min_dim, top + min_dim))

    # Redimensionar a 200×200 px
    img = img.resize((200, 200), Image.LANCZOS)

    buf = BytesIO()
    img.save(buf, format="JPEG", quality=75, optimize=True)
    foto_b64 = "data:image/jpeg;base64," + b64lib.b64encode(buf.getvalue()).decode()

    est.foto_url = foto_b64
    db.commit()
    return {"foto_url": foto_b64}


# ---------------------------------------------------------------------------
# DELETE /estudiantes/{id}/foto  →  Eliminar foto del estudiante
# ---------------------------------------------------------------------------

@router.delete("/{estudiante_id}/foto", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_foto(
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")
    est.foto_url = None
    db.commit()


# ---------------------------------------------------------------------------
# GET /estudiantes/{id}/qr  →  PNG image
# ---------------------------------------------------------------------------

@router.get("/{estudiante_id}/qr")
def obtener_qr(
    estudiante_id: str,
    formato: str = Query("png", description="png | base64"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("i-auxiliar", "p-auxiliar", "s-auxiliar", "admin")),
):
    est = db.query(Estudiante).filter(
        Estudiante.id == estudiante_id, Estudiante.activo == True
    ).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")

    png_bytes = generar_qr_png(
        qr_token=est.qr_token,
        nombre=est.nombre,
        apellido=est.apellido,
        nivel=est.nivel,
        grado=est.grado,
        seccion=est.seccion,
    )

    if formato == "base64":
        b64 = base64.b64encode(png_bytes).decode("utf-8")
        return {
            "estudiante_id": est.id,
            "nombre": f"{est.nombre} {est.apellido}",
            "qr_token": est.qr_token,
            "imagen_base64": f"data:image/png;base64,{b64}",
        }

    if formato == "qr_solo":
        solo_bytes = generar_qr_solo_png(qr_token=est.qr_token)
        b64 = base64.b64encode(solo_bytes).decode("utf-8")
        return {
            "estudiante_id": est.id,
            "qr_token": est.qr_token,
            "imagen_base64": f"data:image/png;base64,{b64}",
        }

    return Response(
        content=png_bytes,
        media_type="image/png",
        headers={
            "Content-Disposition": 'inline; filename="carnet_qr.png"',
        },
    )


# ---------------------------------------------------------------------------
# POST /estudiantes/{id}/qr/regenerar  →  Genera nuevo qr_token
# ---------------------------------------------------------------------------

@router.post("/{estudiante_id}/qr/regenerar", response_model=EstudianteResponse)
def regenerar_qr(
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")

    est.qr_token = generar_qr_token()
    db.commit()
    db.refresh(est)
    return est


# ---------------------------------------------------------------------------
# POST /estudiantes  (admin)
# ---------------------------------------------------------------------------

@router.post("/", response_model=EstudianteResponse, status_code=status.HTTP_201_CREATED)
def crear(
    data: EstudianteCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    # Verificar DNI único
    existing = db.query(Estudiante).filter(Estudiante.dni == data.dni).first()
    if existing:
        if existing.activo:
            raise HTTPException(status.HTTP_409_CONFLICT, f"Ya existe un estudiante con DNI {data.dni}")
        else:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                detail={
                    "codigo": "inactivo",
                    "mensaje": "Este DNI pertenece a un estudiante inactivo",
                    "estudiante": {
                        "id": existing.id,
                        "nombre": existing.nombre,
                        "apellido": existing.apellido,
                        "dni": existing.dni,
                        "nivel": existing.nivel,
                        "grado": existing.grado,
                        "seccion": existing.seccion,
                        "motivo_desactivacion": existing.motivo_desactivacion,
                        "activo": existing.activo,
                    },
                },
            )

    est = Estudiante(
        **data.model_dump(),
        qr_token=generar_qr_token(),
        activo=True,
    )
    db.add(est)
    db.commit()
    db.refresh(est)
    return est


# ---------------------------------------------------------------------------
# PUT /estudiantes/{id}  (admin)
# ---------------------------------------------------------------------------

@router.put("/{estudiante_id}", response_model=EstudianteResponse)
def actualizar(
    estudiante_id: str,
    data: EstudianteUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(est, field, value)

    db.commit()
    db.refresh(est)
    return est


# ---------------------------------------------------------------------------
# POST /estudiantes/{id}/apoderado  →  crear y vincular apoderado
# ---------------------------------------------------------------------------

@router.post("/{estudiante_id}/apoderado", status_code=status.HTTP_201_CREATED)
def vincular_apoderado(
    estudiante_id: str,
    data: dict,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    """
    Crea un usuario apoderado y lo vincula al estudiante.
    Body: {nombre, apellido, dni, email, password}
    Si ya existe un usuario con ese DNI o email, solo crea el vínculo.
    """
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")

    if not data.get("dni"):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Campo requerido: dni")

    # Buscar usuario existente por DNI (cualquier rol — un tutor puede ser padre de un alumno)
    apoderado = db.query(Usuario).filter(Usuario.dni == data["dni"]).first()

    if not apoderado:
        # Apoderado nuevo — validar todos los campos
        for campo in ("nombre", "apellido", "email", "password"):
            if not data.get(campo):
                raise HTTPException(
                    status.HTTP_422_UNPROCESSABLE_ENTITY, f"Campo requerido: {campo}"
                )
        if db.query(Usuario).filter(Usuario.email == data["email"]).first():
            raise HTTPException(status.HTTP_409_CONFLICT, "El email ya está registrado")
        apoderado = Usuario(
            dni=data["dni"],
            nombre=data["nombre"],
            apellido=data["apellido"],
            email=data["email"],
            password_hash=get_password_hash(data["password"]),
            rol="apoderado",
            telefono=data.get("telefono") or None,
            activo=True,
        )
        db.add(apoderado)
        db.flush()

    # Si es personal del colegio (no apoderado puro), marcarle el flag automáticamente
    if apoderado.rol != "apoderado" and not apoderado.es_apoderado:
        apoderado.es_apoderado = True

    # Verificar que no exista ya el vínculo
    vinculo_existente = db.query(ApoderadoEstudiante).filter(
        ApoderadoEstudiante.apoderado_id == apoderado.id,
        ApoderadoEstudiante.estudiante_id == estudiante_id,
    ).first()

    if not vinculo_existente:
        db.add(ApoderadoEstudiante(
            apoderado_id=apoderado.id,
            estudiante_id=estudiante_id,
        ))

    db.commit()
    return {
        "apoderado_id": apoderado.id,
        "nombre": apoderado.nombre,
        "apellido": apoderado.apellido,
        "email": apoderado.email,
        "dni": apoderado.dni,
        "vinculado": True,
    }


# ---------------------------------------------------------------------------
# GET /estudiantes/{id}/apoderados  →  listar apoderados del estudiante
# ---------------------------------------------------------------------------

@router.get("/{estudiante_id}/apoderados")
def listar_apoderados(
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin", "i-auxiliar", "p-auxiliar", "s-auxiliar", "tutor")),
):
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")

    vinculos = db.query(ApoderadoEstudiante).filter(
        ApoderadoEstudiante.estudiante_id == estudiante_id
    ).all()

    result = []
    for v in vinculos:
        apo = db.query(Usuario).filter(Usuario.id == v.apoderado_id).first()
        if apo:
            result.append({
                "id": apo.id,
                "nombre": apo.nombre,
                "apellido": apo.apellido,
                "email": apo.email,
                "dni": apo.dni,
                "telefono": apo.telefono,
                "activo": apo.activo,
            })
    return result


# ---------------------------------------------------------------------------
# DELETE /estudiantes/{id}  →  desactivar (soft delete)
# ---------------------------------------------------------------------------

@router.delete("/{estudiante_id}", status_code=status.HTTP_204_NO_CONTENT)
def desactivar(
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")

    est.activo = False
    db.commit()


# ---------------------------------------------------------------------------
# PATCH /estudiantes/{id}/desactivar  (admin)  →  desactiva con motivo
# ---------------------------------------------------------------------------

@router.patch("/{estudiante_id}/desactivar", response_model=EstudianteResponse)
def desactivar_con_motivo(
    estudiante_id: str,
    data: DesactivarRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    from datetime import datetime, timezone
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")
    if not est.activo:
        raise HTTPException(status.HTTP_409_CONFLICT, "El estudiante ya está desactivado")

    est.activo = False
    est.motivo_desactivacion = data.motivo
    est.desactivado_en = datetime.now(timezone.utc)
    est.desactivado_por = current_user.id
    db.commit()
    db.refresh(est)
    return est


# ---------------------------------------------------------------------------
# PATCH /estudiantes/{id}/reactivar  (admin)  →  reactiva con nueva aula
# ---------------------------------------------------------------------------

@router.patch("/{estudiante_id}/reactivar", response_model=EstudianteResponse)
def reactivar_estudiante(
    estudiante_id: str,
    data: ReactivarRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("admin")),
):
    est = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not est:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Estudiante no encontrado")
    if est.activo:
        raise HTTPException(status.HTTP_409_CONFLICT, "El estudiante ya está activo")

    est.activo = True
    est.nivel = data.nivel
    est.grado = data.grado
    est.seccion = data.seccion
    est.motivo_desactivacion = None
    est.desactivado_en = None
    est.desactivado_por = None
    db.commit()
    db.refresh(est)
    return est
