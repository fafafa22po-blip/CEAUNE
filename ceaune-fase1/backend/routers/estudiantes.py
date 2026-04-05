from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional
import uuid
import io
import openpyxl

from database import get_db
from core.dependencies import require_roles
from core.security import hash_password
from models.estudiante import Estudiante, ApoderadoEstudiante
from models.usuario import Usuario
from schemas.estudiante import EstudianteCreate, EstudianteOut, EstudianteCreado, ApoderadoResumen

router = APIRouter()


@router.get("/", response_model=list[EstudianteOut])
def listar(
    grado: Optional[str] = Query(None),
    seccion: Optional[str] = Query(None),
    activo: bool = Query(default=True),
    search: Optional[str] = Query(None, description="Buscar por nombre, apellido o DNI"),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "auxiliar")),
):
    q = db.query(Estudiante).filter(Estudiante.activo == activo)
    if grado:
        q = q.filter(Estudiante.grado == grado)
    if seccion:
        q = q.filter(Estudiante.seccion == seccion)
    if search:
        term = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Estudiante.nombre.ilike(term),
                Estudiante.apellido.ilike(term),
                Estudiante.dni.ilike(term),
            )
        )
    return q.order_by(Estudiante.apellido, Estudiante.nombre).all()


@router.post("/", response_model=EstudianteCreado, status_code=201)
def crear(
    data: EstudianteCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin")),
):
    if db.query(Estudiante).filter(Estudiante.dni == data.dni).first():
        raise HTTPException(status_code=409, detail="Ya existe un estudiante con ese DNI")

    estudiante = Estudiante(
        id=str(uuid.uuid4()),
        qr_token=str(uuid.uuid4()),
        **data.model_dump(exclude={"apoderado"}),
    )
    db.add(estudiante)
    db.flush()

    apoderado_creado = False
    apoderado_existente = False
    password_temporal = None

    if data.apoderado:
        apoderado = db.query(Usuario).filter(Usuario.dni == data.apoderado.dni).first()

        if apoderado:
            # Apoderado ya existe — solo vincular
            apoderado_existente = True
        else:
            # Verificar que el email no esté en uso por otro usuario
            if db.query(Usuario).filter(Usuario.email == data.apoderado.email).first():
                db.rollback()
                raise HTTPException(
                    status_code=409,
                    detail="El correo del apoderado ya está registrado con otro usuario"
                )

            password_temporal = data.apoderado.dni  # Contraseña temporal = DNI
            apoderado = Usuario(
                id=str(uuid.uuid4()),
                dni=data.apoderado.dni,
                nombre=data.apoderado.nombre,
                apellido=data.apoderado.apellido,
                email=data.apoderado.email,
                password_hash=hash_password(password_temporal),
                rol="apoderado",
            )
            db.add(apoderado)
            db.flush()
            apoderado_creado = True

        # Vincular apoderado ↔ estudiante (si no existe ya el vínculo)
        existe_vinculo = db.query(ApoderadoEstudiante).filter(
            ApoderadoEstudiante.apoderado_id == apoderado.id,
            ApoderadoEstudiante.estudiante_id == estudiante.id,
        ).first()

        if not existe_vinculo:
            db.add(ApoderadoEstudiante(
                id=str(uuid.uuid4()),
                apoderado_id=apoderado.id,
                estudiante_id=estudiante.id,
            ))

    db.commit()
    db.refresh(estudiante)

    return EstudianteCreado(
        estudiante=EstudianteOut.model_validate(estudiante),
        apoderado_creado=apoderado_creado,
        apoderado_existente=apoderado_existente,
        password_temporal=password_temporal,
    )


@router.get("/{estudiante_id}", response_model=EstudianteOut)
def obtener(
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "auxiliar")),
):
    e = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    return e


@router.get("/{estudiante_id}/apoderados", response_model=list[ApoderadoResumen])
def apoderados_de_estudiante(
    estudiante_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "auxiliar")),
):
    e = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")

    rels = db.query(ApoderadoEstudiante).filter(
        ApoderadoEstudiante.estudiante_id == estudiante_id
    ).all()

    apoderados = []
    for rel in rels:
        u = db.query(Usuario).filter(
            Usuario.id == rel.apoderado_id,
            Usuario.activo == True,
        ).first()
        if u:
            apoderados.append(u)
    return apoderados


@router.get("/{estudiante_id}/qr")
def get_qr(
    estudiante_id: str,
    solo_qr: bool = Query(default=False, description="True = solo el QR sin tarjeta"),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin", "auxiliar")),
):
    e = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")

    from services.qr_service import generar_qr_bytes, generar_qr_card_bytes

    if solo_qr:
        img_bytes = generar_qr_bytes(e.qr_token)
    else:
        img_bytes = generar_qr_card_bytes(
            qr_token=e.qr_token,
            nombre=e.nombre,
            apellido=e.apellido,
            grado=e.grado,
            seccion=e.seccion,
            dni=e.dni,
        )

    return StreamingResponse(
        io.BytesIO(img_bytes),
        media_type="image/png",
        headers={"Content-Disposition": f'inline; filename="qr_{e.dni}.png"'},
    )


@router.post("/importar-excel")
def importar_excel(
    file: UploadFile = File(..., description="Excel .xlsx con columnas: dni, nombre, apellido, grado, seccion"),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("admin")),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {e}")

    COLUMNAS = {"dni", "nombre", "apellido", "grado", "seccion"}
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    faltantes = COLUMNAS - set(headers)
    if faltantes:
        raise HTTPException(status_code=400, detail=f"Columnas faltantes en el Excel: {faltantes}")

    idx = {col: headers.index(col) for col in COLUMNAS}

    creados = []
    omitidos = []
    errores = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            dni      = str(row[idx["dni"]]).strip().zfill(8) if row[idx["dni"]] else None
            nombre   = str(row[idx["nombre"]]).strip()   if row[idx["nombre"]]   else None
            apellido = str(row[idx["apellido"]]).strip() if row[idx["apellido"]] else None
            grado    = str(row[idx["grado"]]).strip()    if row[idx["grado"]]    else None
            seccion  = str(row[idx["seccion"]]).strip().upper() if row[idx["seccion"]] else None

            if not all([dni, nombre, apellido, grado, seccion]):
                errores.append({"fila": row_num, "error": "Campos incompletos"})
                continue

            if db.query(Estudiante).filter(Estudiante.dni == dni).first():
                omitidos.append({"fila": row_num, "dni": dni, "razon": "DNI duplicado"})
                continue

            e = Estudiante(
                id=str(uuid.uuid4()),
                dni=dni,
                nombre=nombre,
                apellido=apellido,
                grado=grado,
                seccion=seccion,
                qr_token=str(uuid.uuid4()),
            )
            db.add(e)
            creados.append({"fila": row_num, "dni": dni, "nombre": f"{nombre} {apellido}"})

        except Exception as exc:
            errores.append({"fila": row_num, "error": str(exc)})

    db.commit()

    return {
        "creados": len(creados),
        "omitidos": len(omitidos),
        "errores": len(errores),
        "detalle_creados": creados,
        "detalle_omitidos": omitidos,
        "detalle_errores": errores,
    }
